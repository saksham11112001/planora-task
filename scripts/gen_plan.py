from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side, GradientFill)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Helpers ────────────────────────────────────────────────────────────────
def cell(ws, row, col, value='', bold=False, italic=False, size=10,
         fg=None, bg=None, align='left', wrap=False, border=False, color='000000'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Arial', bold=bold, italic=italic, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if bg:
        c.fill = PatternFill('solid', start_color=bg)
    if border:
        thin = Side(style='thin', color='D0D0D0')
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def header_row(ws, row, labels, widths, bg='1A3C5E', fg='FFFFFF', height=22):
    ws.row_dimensions[row].height = height
    for col, (label, w) in enumerate(zip(labels, widths), 1):
        c = cell(ws, row, col, label, bold=True, size=10, bg=bg, align='center', border=True, color=fg)
        ws.column_dimensions[get_column_letter(col)].width = w

def merge_title(ws, row, col1, col2, text, bg='1A3C5E', fg='FFFFFF', size=13):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.font = Font(name='Arial', bold=True, size=size, color=fg)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28

PRIORITY_BG = {'P0 Critical': 'FFD6D6', 'P1 High': 'FFE8CC', 'P2 Medium': 'FFF9CC', 'P3 Low': 'E8F5E9'}
PRIORITY_FG = {'P0 Critical': 'C0392B', 'P1 High': 'D35400', 'P2 Medium': 'B7950B', 'P3 Low': '196F3D'}
STATUS_BG   = {'✅ Done': 'D5F5E3', '🔄 Partial': 'FEF9E7', '⏳ Pending': 'EBF5FB', '🔮 Future': 'F4ECF7'}
WEEK_COLORS = ['D6EAF8', 'D5F5E3', 'FEF9E7', 'FDEDEC', 'F4ECF7']

# ══════════════════════════════════════════════════════════════════════════
# SHEET 1 — SCALABILITY FIXES TRACKER
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '🔧 Fixes Tracker'
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = 'A4'

merge_title(ws1, 1, 1, 7, 'Taska — Scalability & Security Fixes Tracker', '0D4C73', 'FFFFFF', 14)
merge_title(ws1, 2, 1, 7, 'All fixes implemented before May 15 Launch  |  Last updated: Apr 2026', '1A6B96', 'FFFFFF', 10)

cols  = ['#', 'Fix / Issue', 'Priority', 'File(s) Changed', 'What Was Fixed', 'Status', 'Notes']
widths= [4, 36, 13, 38, 46, 12, 30]
header_row(ws1, 3, cols, widths, bg='0D4C73')

fixes = [
    (1,  'Rate limiting on all API routes',              'P0 Critical', 'middleware.ts\nlib/utils/rateLimit.ts (NEW)',
          'Sliding-window rate limiter: 120 req/min general, 10/5min auth, 20/min upload, 5/5min import, 10/hr reports',
          '✅ Done', 'In-memory per instance. Add Upstash Redis for distributed limiting at 1000+ users.'),
    (2,  'Razorpay webhook: dangerous "business" default','P0 Critical', 'app/api/webhooks/razorpay/route.ts',
          'Unknown plan_id now logs error & keeps current tier instead of silently upgrading to Business. DB error returns 500 so Razorpay retries.',
          '✅ Done', 'Set RAZORPAY_BUSINESS_PLAN_ID env var to cover all 3 plan IDs.'),
    (3,  'Upload memory bomb (OOM risk)',                 'P0 Critical', 'app/api/portal/[token]/upload/route.ts',
          'Content-Length header checked BEFORE req.formData() so 500 MB files never load into RAM.',
          '✅ Done', 'Also double-checks file.size after parse for spoofed Content-Length.'),
    (4,  'Portal N+1: fetching all org instances in JS',  'P0 Critical', 'app/api/portal/[token]/route.ts',
          'ca_task_instances now filtered by client assignment IDs at DB level (.in() query). Added .limit(200) & .limit(100) on history.',
          '✅ Done', 'Was fetching 50k+ rows for entire org, then filtering in JS. Now fetches only this client\'s rows.'),
    (5,  'Duplicate contradictory plan limits',           'P1 High',     'app/api/team/route.ts',
          'Removed hardcoded local memberLimit() / effectivePlan() functions. Now imports from lib/utils/planGate.ts.',
          '✅ Done', 'Was: free=3, starter=10. Correct: free=5, starter=15. Trial logic now uses planGate expiry check.'),
    (6,  'Inngest: unbounded due-soon task fetch',        'P1 High',     'lib/inngest/functions/dailyReminders.ts',
          'Added .limit(500) to due-soon fetch step and .limit(300) to escalation step.',
          '✅ Done', 'Prevents 10,000-step Inngest jobs. Jobs over the limit will process next day.'),
    (7,  'WhatsApp: no retry on failure',                 'P1 High',     'lib/whatsapp/client.ts',
          'Added exponential backoff retry (3 attempts, 500ms / 1000ms delays). AbortSignal.timeout(10s) per attempt. Terminal errors (invalid template) skip retry.',
          '✅ Done', 'Messages previously lost silently on any network hiccup.'),
    (8,  'No maxDuration on heavy routes',                'P1 High',     'app/api/reports/export/route.ts\napp/api/ca/assignments/route.ts\napp/api/import/route.ts\napp/api/portal/[token]/route.ts\napp/api/portal/[token]/upload/route.ts',
          'Added export const maxDuration: 60s (reports, portal), 300s (import), 30s (ca/assignments, portal GET).',
          '✅ Done', 'Without this, Vercel kills requests at 10s default on hobby or 60s on Pro.'),
    (9,  'Supabase: not using connection pooler',         'P2 Medium',   'lib/supabase/admin.ts\nlib/supabase/server.ts',
          'admin.ts now uses SUPABASE_POOLER_URL env var if set (pgBouncer transaction mode, port 6543). Falls back to direct URL.',
          '✅ Done', 'ACTION: Add SUPABASE_POOLER_URL to Vercel env vars. Find it in Supabase → Settings → Database → Connection Pooling.'),
    (10, 'No Cache-Control on GET endpoints',             'P2 Medium',   'app/api/tasks/route.ts\napp/api/projects/route.ts\napp/api/clients/route.ts\napp/api/team/route.ts',
          'Added "Cache-Control: private, max-age=30, stale-while-revalidate=60" on all major list GETs.',
          '✅ Done', 'Reduces repeat DB hits from same browser. 30s stale-while-revalidate for near-instant repeat loads.'),
    (11, 'report-issue: no size guard',                   'P2 Medium',   'app/api/report-issue/route.ts',
          'Content-Length checked (10 MB max) before parsing FormData body.',
          '✅ Done', ''),
    (12, 'Portal: history query no DB filter',            'P2 Medium',   'app/api/portal/[token]/route.ts',
          'History query now uses .in(assignment_id, clientAssignmentIds) filter at DB level + .limit(100).',
          '✅ Done', 'Part of the portal N+1 fix above.'),
    (13, 'Verify DB composite indexes exist',             'P2 Medium',   'Supabase Dashboard (SQL editor)',
          'Run index verification queries in Supabase SQL editor (see Notes column).',
          '⏳ Pending', 'Run: CREATE INDEX IF NOT EXISTS idx_org_members_user_active ON org_members(user_id, is_active);\nCREATE INDEX IF NOT EXISTS idx_tasks_org_due ON tasks(org_id, due_date) WHERE is_archived = false;\nCREATE INDEX IF NOT EXISTS idx_ca_instances_assignment ON ca_task_instances(assignment_id, due_date);'),
    (14, 'Add Upstash Redis for distributed rate limiting','P3 Low',     'middleware.ts + lib/utils/rateLimit.ts',
          'Current in-memory rate limiter works per Vercel instance. For 1000+ users across instances, replace with Upstash ratelimit.',
          '🔮 Future', 'npm i @upstash/ratelimit @upstash/redis\nAdd UPSTASH_REDIS_REST_URL + UPSTASH_REDIS_REST_TOKEN to env vars.'),
    (15, 'Silent catch{} blocks in 8+ routes',            'P3 Low',     'Various API routes',
          'Silent empty catch blocks swallow errors. Should at minimum log them.',
          '🔮 Future', 'Low risk but reduces observability. Add console.error() or Sentry in future sprint.'),
]

row = 4
for f in fixes:
    idx, title, prio, files, what, status, notes = f
    ws1.row_dimensions[row].height = 55
    bg_p = PRIORITY_BG.get(prio, 'FFFFFF')
    fg_p = PRIORITY_FG.get(prio, '000000')
    bg_s = STATUS_BG.get(status, 'FFFFFF')
    bg_alt = 'F8FAFB' if row % 2 == 0 else 'FFFFFF'

    cell(ws1, row, 1, idx,    bold=True,  align='center', bg=bg_alt, border=True)
    cell(ws1, row, 2, title,  bold=True,  bg=bg_alt, border=True, wrap=True)
    c = cell(ws1, row, 3, prio, bold=True, align='center', bg=bg_p, border=True, color=fg_p)
    cell(ws1, row, 4, files,  italic=True, size=9, bg=bg_alt, border=True, wrap=True)
    cell(ws1, row, 5, what,   bg=bg_alt, border=True, wrap=True, size=9)
    cell(ws1, row, 6, status, bold=True, align='center', bg=bg_s, border=True)
    cell(ws1, row, 7, notes,  italic=True, size=9, bg=bg_alt, border=True, wrap=True)
    row += 1

# ══════════════════════════════════════════════════════════════════════════
# SHEET 2 — LAUNCH TIMELINE
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('📅 Launch Timeline')
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = 'A4'

merge_title(ws2, 1, 1, 7, 'Taska — May 15 Launch Timeline (Apr 15 → May 15, 2026)', '1B4F72', 'FFFFFF', 14)
merge_title(ws2, 2, 1, 7, 'Official launch on 15 May 2026  |  30-day plan', '21618C', 'FFFFFF', 10)

cols2  = ['Week', 'Dates', 'Theme', 'Task', 'Owner', 'Priority', 'Status']
widths2= [8, 18, 22, 54, 16, 13, 12]
header_row(ws2, 3, cols2, widths2, bg='1B4F72')

timeline = [
    # Week 1
    (1,'Apr 15–21','🔧 Critical Tech Fixes','✅ Rate limiting added to middleware (done)','Dev','P0 Critical','✅ Done'),
    (1,'Apr 15–21','🔧 Critical Tech Fixes','✅ Razorpay webhook plan-tier safety (done)','Dev','P0 Critical','✅ Done'),
    (1,'Apr 15–21','🔧 Critical Tech Fixes','✅ Upload memory bomb fix (done)','Dev','P0 Critical','✅ Done'),
    (1,'Apr 15–21','🔧 Critical Tech Fixes','✅ Portal N+1 query fixed (done)','Dev','P0 Critical','✅ Done'),
    (1,'Apr 15–21','🔧 Critical Tech Fixes','Add SUPABASE_POOLER_URL to Vercel env vars','Dev','P1 High','⏳ Pending'),
    (1,'Apr 15–21','🔧 Critical Tech Fixes','Run DB index creation SQL in Supabase dashboard','Dev','P2 Medium','⏳ Pending'),
    (1,'Apr 15–21','🔧 Critical Tech Fixes','Register WhatsApp templates with MSG91 (takes 24–48h approval)','Dev','P1 High','⏳ Pending'),
    (1,'Apr 15–21','🔧 Critical Tech Fixes','Set all prod env vars: Razorpay live keys, Resend, MSG91','Dev','P0 Critical','⏳ Pending'),
    # Week 2
    (2,'Apr 22–28','💳 Payments & Portal','Create 6 Razorpay plan IDs (3 tiers × monthly + annual)','Founder','P0 Critical','⏳ Pending'),
    (2,'Apr 22–28','💳 Payments & Portal','E2E test: signup → trial → upgrade → subscription active','Dev','P0 Critical','⏳ Pending'),
    (2,'Apr 22–28','💳 Payments & Portal','Test Razorpay webhook with ngrok (cancelled, expired, charged)','Dev','P0 Critical','⏳ Pending'),
    (2,'Apr 22–28','💳 Payments & Portal','Partner review: client portal view — confirm UX for CA client','Partner','P1 High','⏳ Pending'),
    (2,'Apr 22–28','💳 Payments & Portal','Adjust portal UI based on partner feedback','Dev','P1 High','⏳ Pending'),
    (2,'Apr 22–28','💳 Payments & Portal','Test WhatsApp: task assigned, approval needed, due-soon messages','Dev','P1 High','⏳ Pending'),
    (2,'Apr 22–28','💳 Payments & Portal','Load test portal token endpoint — verify N+1 fix works at scale','Dev','P2 Medium','⏳ Pending'),
    # Week 3
    (3,'Apr 29–May 5','🎬 Content & Marketing','Record demo video: full user journey (signup→task→recurring→approval)','Founder','P1 High','⏳ Pending'),
    (3,'Apr 29–May 5','🎬 Content & Marketing','Record CA compliance demo: spawn, portal, document upload','Founder','P1 High','⏳ Pending'),
    (3,'Apr 29-May 5','🎬 Content & Marketing','Record 60-second promo: India CA task manager value prop','Founder','P1 High','⏳ Pending'),
    (3,'Apr 29–May 5','🎬 Content & Marketing','Set up landing page: pricing, screenshots, demo embed','Dev','P2 Medium','⏳ Pending'),
    (3,'Apr 29–May 5','🎬 Content & Marketing','Write LinkedIn post, Twitter thread, WhatsApp broadcast copy','Marketing','P2 Medium','⏳ Pending'),
    (3,'Apr 29–May 5','🎬 Content & Marketing','Create Product Hunt draft listing (schedule for May 15)','Founder','P2 Medium','⏳ Pending'),
    (3,'Apr 29–May 5','🎬 Content & Marketing','Set up issue/feedback admin view (Supabase issue_reports table or admin page)','Dev','P3 Low','⏳ Pending'),
    # Week 4
    (4,'May 6–12','🧪 Beta & Polish','Beta test with 5–10 real CA firms — collect feedback','Founder','P0 Critical','⏳ Pending'),
    (4,'May 6–12','🧪 Beta & Polish','Fix critical bugs from beta feedback','Dev','P0 Critical','⏳ Pending'),
    (4,'May 6–12','🧪 Beta & Polish','Full E2E regression: all major user flows','Dev','P1 High','⏳ Pending'),
    (4,'May 6–12','🧪 Beta & Polish','Verify all Inngest jobs fire correctly in production','Dev','P1 High','⏳ Pending'),
    (4,'May 6–12','🧪 Beta & Polish','Prepare launch day playbook: monitoring plan, escalation contacts','Founder','P1 High','⏳ Pending'),
    (4,'May 6–12','🧪 Beta & Polish','Soft launch to waitlist/beta users on May 14','Founder','P1 High','⏳ Pending'),
    # Launch day
    (5,'May 15','🚀 Launch Day','Product Hunt launch (schedule 12:01 AM PST, mobilise network for upvotes)','Founder','P0 Critical','⏳ Pending'),
    (5,'May 15','🚀 Launch Day','LinkedIn long-form post: "We built a task manager for CA firms"','Founder','P0 Critical','⏳ Pending'),
    (5,'May 15','🚀 Launch Day','Email blast to waitlist with demo video link','Founder','P1 High','⏳ Pending'),
    (5,'May 15','🚀 Launch Day','Monitor Vercel dashboard + Supabase for any errors/overload','Dev','P0 Critical','⏳ Pending'),
    (5,'May 15','🚀 Launch Day','WhatsApp broadcast to CA network with app link','Founder','P1 High','⏳ Pending'),
]

row = 4
prev_week = None
week_bg_map = {1:'D6EAF8', 2:'D5F5E3', 3:'FEF9E7', 4:'FDEDEC', 5:'F4ECF7'}
for t in timeline:
    wk, dates, theme, task, owner, prio, status = t
    ws2.row_dimensions[row].height = 28
    bg = week_bg_map.get(wk, 'FFFFFF')
    bg_p = PRIORITY_BG.get(prio, 'FFFFFF')
    fg_p = PRIORITY_FG.get(prio, '000000')
    bg_s = STATUS_BG.get(status, 'FFFFFF')

    cell(ws2, row, 1, f'W{wk}' if wk < 5 else 'GO!', bold=True, align='center', bg=bg, border=True)
    cell(ws2, row, 2, dates, bg=bg, border=True, size=9)
    cell(ws2, row, 3, theme, bold=True, bg=bg, border=True, size=9)
    cell(ws2, row, 4, task, bg='FFFFFF' if task.startswith('✅') else 'FFFFFF', border=True, wrap=True)
    cell(ws2, row, 5, owner, align='center', bg='F8F9FA', border=True, size=9)
    cell(ws2, row, 6, prio, bold=True, align='center', bg=bg_p, border=True, color=fg_p, size=9)
    cell(ws2, row, 7, status, bold=True, align='center', bg=bg_s, border=True, size=9)
    row += 1

# ══════════════════════════════════════════════════════════════════════════
# SHEET 3 — MARKETING PLAN
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('📣 Marketing Plan')
ws3.sheet_view.showGridLines = False

merge_title(ws3, 1, 1, 6, 'Taska — Go-To-Market & Marketing Plan', '1B4F72', 'FFFFFF', 14)
merge_title(ws3, 2, 1, 6, 'Target: CA Firms & Independent CAs in India', '21618C', 'FFFFFF', 10)

# Phase sections
sections = [
    ('PRE-LAUNCH (Now → May 14)', '1A5276', [
        ('LinkedIn — Building in Public','Weekly posts: progress, features, behind the scenes. Target: CA India groups.','Founder','Free','Ongoing','Build audience before launch'),
        ('WhatsApp Broadcast','Personal network of CA contacts. Share early access link & demo video.','Founder','Free','Apr 20+','Warm leads convert best'),
        ('CA Community Outreach','Post in ICAI alumni groups, CA club WhatsApp groups, CA coaching centre networks.','Founder','Free','Apr 22+','High-intent audience'),
        ('Early Access Waitlist','Simple form → email sequence via Resend. Offer 3-month free Pro trial.','Dev + Founder','Free','Live now','Collect leads before launch'),
        ('Product Hunt Draft','Schedule PH listing for May 15 12:01 AM PST. Get 50+ upvote commitments from network.','Founder','Free','May 10','PH launch = free traffic spike'),
    ]),
    ('LAUNCH DAY (May 15)', '117A65', [
        ('Product Hunt Launch','Go live on PH. Post in all communities asking for upvotes. Aim for Top 5 of the day.','Founder','Free','May 15','Critical for early traction'),
        ('LinkedIn Long-Form Post','"We built a task manager for CA firms — here is why" story post with demo video.','Founder','Free','May 15 9 AM','High-engagement format'),
        ('Email Blast to Waitlist','Announce launch with demo video, pricing, free trial link.','Founder','Resend (free tier)','May 15','Direct to interested users'),
        ('Twitter/X Thread','Thread: "5 problems every CA firm faces with task management + how we solved them"','Founder','Free','May 15','Broader tech audience'),
        ('WhatsApp Blast to CA Network','Personal message to all CA contacts with app link and 1-min promo video.','Founder','Free','May 15','Highest conversion channel'),
    ]),
    ('POST-LAUNCH (May 16+)', '6E2F9E', [
        ('Customer Success Follow-ups','Weekly check-ins with early users → collect testimonials and case studies.','Founder','Free','Ongoing','Fuel word of mouth'),
        ('Case Study: First CA Firm','Quantify: "Reduced missed deadlines by X%" — publish as blog + LinkedIn post.','Founder','Free','Jun 2026','Social proof for sales'),
        ('Google/Meta Ads (Optional)','Retargeting visitors + CA-profession targeting. Start small (₹5k/mo test budget).','Marketing','₹5k–₹20k/mo','May 20+','Only after organic validated'),
        ('CA College/Coaching Outreach','Approach CA Final coaching centres — offer student discount, generate word of mouth.','Founder','Free','Jun 2026','Long-term pipeline'),
        ('YouTube Tutorial Videos','How-to videos for CA compliance module, recurring tasks, client portal.','Founder','Free','Jun 2026','SEO + trust building'),
        ('Referral Program','30-day free Pro for every paid user you refer. Built into billing page.','Dev','Build cost','Jun 2026','Viral growth mechanism'),
    ]),
]

row = 4
cols3  = ['Activity', 'Description', 'Owner', 'Cost', 'Timeline', 'Why It Matters']
widths3= [28, 52, 12, 18, 14, 32]
header_row(ws3, row, cols3, widths3, bg='1B4F72')
row += 1

for sec_title, sec_color, items in sections:
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws3.cell(row=row, column=1, value=sec_title)
    c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill = PatternFill('solid', start_color=sec_color)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws3.row_dimensions[row].height = 22
    row += 1

    for item in items:
        act, desc, owner, cost, timeline, why = item
        ws3.row_dimensions[row].height = 36
        alt = 'F8FAFB' if row % 2 == 0 else 'FFFFFF'
        cell(ws3, row, 1, act,      bold=True, bg=alt, border=True, wrap=True)
        cell(ws3, row, 2, desc,     bg=alt, border=True, wrap=True, size=9)
        cell(ws3, row, 3, owner,    align='center', bg=alt, border=True, size=9)
        cell(ws3, row, 4, cost,     align='center', bg=alt, border=True, size=9)
        cell(ws3, row, 5, timeline, align='center', bg=alt, border=True, size=9)
        cell(ws3, row, 6, why,      italic=True, bg=alt, border=True, wrap=True, size=9)
        row += 1

# ══════════════════════════════════════════════════════════════════════════
# SHEET 4 — ENV VARS CHECKLIST
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('⚙️ Env Vars Checklist')
ws4.sheet_view.showGridLines = False

merge_title(ws4, 1, 1, 5, 'Taska — Production Environment Variables Checklist', '1B4F72', 'FFFFFF', 14)
merge_title(ws4, 2, 1, 5, 'All must be set in Vercel → Project → Settings → Environment Variables before launch', '21618C', 'FFFFFF', 10)

cols4   = ['Variable', 'Required?', 'Where to Get It', 'Status', 'Notes']
widths4 = [38, 12, 46, 12, 34]
header_row(ws4, 3, cols4, widths4, bg='1B4F72')

env_vars = [
    # Supabase
    ('── SUPABASE ──', '', '', '', ''),
    ('NEXT_PUBLIC_SUPABASE_URL',    '✅ Required', 'Supabase Dashboard → Settings → API', '⏳ Set?', ''),
    ('NEXT_PUBLIC_SUPABASE_ANON_KEY','✅ Required', 'Supabase Dashboard → Settings → API', '⏳ Set?', ''),
    ('SUPABASE_SERVICE_ROLE_KEY',   '✅ Required', 'Supabase Dashboard → Settings → API', '⏳ Set?', 'Keep secret — server only'),
    ('SUPABASE_POOLER_URL',         '🔶 Recommended', 'Supabase → Settings → Database → Connection Pooling (port 6543)', '⏳ Set?', 'Prevents "too many connections" at scale'),
    # App
    ('── APP ──', '', '', '', ''),
    ('NEXT_PUBLIC_APP_URL',         '✅ Required', 'Your production domain e.g. https://app.taska.in', '⏳ Set?', ''),
    # Razorpay
    ('── RAZORPAY ──', '', '', '', ''),
    ('RAZORPAY_KEY_ID',             '✅ Required', 'Razorpay Dashboard → Settings → API Keys (use LIVE key)', '⏳ Set?', 'Must be rzp_live_... not rzp_test_...'),
    ('RAZORPAY_KEY_SECRET',         '✅ Required', 'Razorpay Dashboard → Settings → API Keys', '⏳ Set?', ''),
    ('RAZORPAY_WEBHOOK_SECRET',     '✅ Required', 'Razorpay Dashboard → Webhooks → create webhook → copy secret', '⏳ Set?', ''),
    ('RAZORPAY_STARTER_PLAN_ID',    '✅ Required', 'Razorpay Dashboard → Subscriptions → Plans → Starter plan ID', '⏳ Set?', 'Format: plan_XXXXXXXX'),
    ('RAZORPAY_PRO_PLAN_ID',        '✅ Required', 'Razorpay Dashboard → Subscriptions → Plans → Pro plan ID', '⏳ Set?', ''),
    ('RAZORPAY_BUSINESS_PLAN_ID',   '✅ Required', 'Razorpay Dashboard → Subscriptions → Plans → Business plan ID', '⏳ Set?', 'NEW — needed for webhook fix'),
    # Email
    ('── EMAIL (RESEND) ──', '', '', '', ''),
    ('RESEND_API_KEY',              '✅ Required', 'resend.com → API Keys', '⏳ Set?', ''),
    ('FROM_EMAIL',                  '✅ Required', 'e.g. Taska <noreply@taska.in>', '⏳ Set?', 'Must be a verified domain in Resend'),
    # Inngest
    ('── INNGEST ──', '', '', '', ''),
    ('INNGEST_EVENT_KEY',           '✅ Required', 'app.inngest.com → your app → Event Keys', '⏳ Set?', ''),
    ('INNGEST_SIGNING_KEY',         '✅ Required', 'app.inngest.com → your app → Signing Keys', '⏳ Set?', ''),
    # WhatsApp
    ('── WHATSAPP (MSG91) ──', '', '', '', ''),
    ('MSG91_AUTH_KEY',              '🔶 Optional', 'msg91.com → API → Auth Key', '⏳ Set?', 'Required for WhatsApp notifications'),
    ('MSG91_WHATSAPP_SENDER_ID',    '🔶 Optional', 'msg91.com → WhatsApp → Sender', '⏳ Set?', 'Required for WhatsApp notifications'),
]

row = 4
for ev in env_vars:
    var, req, where, status, notes = ev
    ws4.row_dimensions[row].height = 28
    is_section = var.startswith('──')
    if is_section:
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws4.cell(row=row, column=1, value=var)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='2C3E50')
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    else:
        alt = 'F8FAFB' if row % 2 == 0 else 'FFFFFF'
        req_bg = 'D5F5E3' if req == '✅ Required' else 'FEF9E7'
        req_fg = '1E8449' if req == '✅ Required' else 'D35400'
        cell(ws4, row, 1, var,    bold=True, bg=alt, border=True, size=9)
        c = cell(ws4, row, 2, req, bold=True, align='center', bg=req_bg, border=True, size=9, color=req_fg)
        cell(ws4, row, 3, where,  bg=alt, border=True, wrap=True, size=9, italic=True)
        cell(ws4, row, 4, status, align='center', bg='FEF9E7', border=True, size=9)
        cell(ws4, row, 5, notes,  bg=alt, border=True, wrap=True, size=9)
    row += 1

# ══════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════
out = r'C:\Users\saksh\Desktop\Taska_Launch_Plan.xlsx'
wb.save(out)
print('Saved:', out)
