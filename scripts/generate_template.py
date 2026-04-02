#!/usr/bin/env python3
import sys, json, base64, tempfile, os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

data = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
member_emails = data.get('memberEmails', ['alex@yourcompany.com','priya@yourcompany.com','sam@yourcompany.com'])
member_names  = data.get('memberNames',  ['Alex Johnson','Priya Sharma','Sam Gupta'])
member_roles  = data.get('memberRoles',  ['manager','member','member'])
client_names  = data.get('clientNames',  ['Acme Corp','Garg Sons','Mehra & Co'])
client_rows   = data.get('clientRows',   [
    ['Acme Corp','hello@acme.com','+91 9876543210','Acme Corp Ltd','Technology','active','#6366f1'],
    ['Garg Sons','accounts@gargsons.com','+91 9988776655','Garg Sons Ltd','Retail','active','#ea580c'],
    ['Mehra & Co','info@mehraandco.com','','Mehra & Co CA','Finance','active','#0d9488'],
])
ca_mode = data.get('caMode', False)

wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill('solid', fgColor='0F172A')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
NORMAL_FONT = Font(size=10)
TEAL_FILL   = PatternFill('solid', fgColor='F0FDFA')
GREY_FILL   = PatternFill('solid', fgColor='F8FAFC')

def make_sheet(wb, name, headers, widths, rows, validations):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22
    for r, row in enumerate(rows, 2):
        fill = TEAL_FILL if r % 2 == 0 else GREY_FILL
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=str(val) if val is not None else '')
            cell.fill = fill
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[r].height = 18
    for col_letter, formula in validations:
        dv = DataValidation(type='list', formula1=formula, allow_blank=True, showDropDown=False)
        dv.sqref = f'{col_letter}2:{col_letter}1000'
        dv.showErrorMessage = True
        dv.showInputMessage = True
        dv.prompt = 'Select from the list'
        dv.promptTitle = 'Options'
        ws.add_data_validation(dv)
    return ws

e_f     = '"' + ','.join(member_emails) + '"'
c_f     = '"' + ','.join(client_names)  + '"'
role_f  = '"owner,admin,manager,member,viewer"'
prio_f  = '"none,low,medium,high,urgent"'
stat_f  = '"todo,in_progress,completed,blocked"'
freq_f  = '"daily,weekly,bi_weekly,monthly,quarterly,annual"'
cst_f   = '"active,inactive,lead"'
pst_f   = '"active,on_hold,completed,cancelled"'

ws_r = wb.create_sheet('=README')
ws_r.title = '\U0001F4D6 README'
ws_r.sheet_view.showGridLines = False
ws_r.column_dimensions['A'].width = 82
readme_lines = [
    ('PLANORA IMPORT TEMPLATE — personalised for your workspace', Font(bold=True,size=14,color='0D9488'), PatternFill('solid',fgColor='F0FDFA')),
    ('', None, None),
    ('FILL IN ORDER:', Font(bold=True,size=11), None),
    ('  Step 1 \u2192  \U0001F465 Team Members  \u2014 team emails become Assignee & Approver dropdowns', Font(size=10), None),
    ('  Step 2 \u2192  \U0001F3E2 Clients       \u2014 client names become Client Name dropdowns', Font(size=10), None),
    ('  Step 3 \u2192  \U0001F4C1 Projects      \u2014 reference client names from Step 2', Font(size=10), None),
    ('  Step 4 \u2192  Tasks / One-Time Tasks / Recurring Tasks', Font(size=10), None),
    ('', None, None),
    ('DROPDOWN COLUMNS \u2014 click any cell to select from list:', Font(bold=True,size=11), None),
    (f'  Assignee Email  \u2192  {len(member_emails)} member(s) from your workspace', Font(size=10,color='0D9488'), None),
    (f'  Approver Email  \u2192  {len(member_emails)} member(s) from your workspace', Font(size=10,color='0D9488'), None),
    (f'  Client Name     \u2192  {len(client_names)} client(s) from your workspace', Font(size=10,color='0D9488'), None),
    ('  Role            \u2192  owner | admin | manager | member | viewer', Font(size=10), None),
    ('  Priority        \u2192  none | low | medium | high | urgent', Font(size=10), None),
    ('  Status          \u2192  todo | in_progress | completed | blocked', Font(size=10), None),
    ('  Frequency       \u2192  daily | weekly | bi_weekly | monthly | quarterly | annual', Font(size=10), None),
    ('', None, None),
    ('MULTI-ASSIGNEE: comma-separate emails  e.g.  alex@co.com,priya@co.com', Font(size=10,italic=True,color='64748B'), None),
    ('DATES: YYYY-MM-DD format  e.g.  2025-08-31', Font(size=10,italic=True,color='64748B'), None),
    ('Columns marked * are required. Blank rows are skipped on import.', Font(bold=True,size=10,color='DC2626'), PatternFill('solid',fgColor='FEF2F2')),
]
for i, (text, font, fill) in enumerate(readme_lines, 1):
    cell = ws_r.cell(row=i, column=1, value=text)
    if font: cell.font = font
    if fill: cell.fill = fill
    cell.alignment = Alignment(vertical='center')
    ws_r.row_dimensions[i].height = 20

mem_data = list(zip(member_names, member_emails, member_roles, ['']*len(member_names)))
make_sheet(wb, '\U0001F465 Team Members', ['Full Name *','Email *','Role *','Notes'], [25,32,14,25],
    [list(r) for r in mem_data], [('C', role_f)])

make_sheet(wb, '\U0001F3E2 Clients', ['Client Name *','Contact Email','Phone','Company','Industry','Status','Color','Notes'],
    [24,28,17,24,14,12,10,22], [r[:8] + [''] * (8 - len(r[:8])) for r in client_rows], [('F', cst_f)])

p1e = member_emails[0] if member_emails else ''
p2e = member_emails[1] if len(member_emails)>1 else p1e
c1  = client_names[0]  if client_names else ''
c2  = client_names[1]  if len(client_names)>1 else c1
make_sheet(wb, '\U0001F4C1 Projects',
    ['Project Name *','Color','Status','Due Date','Owner Email','Client Name','Hours Budget','Description'],
    [24,10,14,14,32,22,13,28],
    [['Project Alpha','#6366f1','active','2025-08-31',p1e,c1,'40',''],
     ['Project Beta', '#ea580c','active','2025-09-30',p2e,c2,'',  '']],
    [('C',pst_f),('E',e_f),('F',c_f)])

t1e = member_emails[0] if member_emails else ''
t2e = member_emails[1] if len(member_emails)>1 else t1e
make_sheet(wb, '\u2705 Tasks',
    ['Task Title *','Project Name','Assignee Email(s)','Approver Email','Priority','Due Date','Status','Client Name','Est. Hours','Description'],
    [28,20,32,32,10,14,14,20,10,28],
    [['Design wireframes','Project Alpha',t1e,p1e,'high','2025-07-15','todo',c1,'4',''],
     ['Review documents','Project Beta',t2e,p2e,'medium','2025-07-20','todo',c2,'2','']],
    [('C',e_f),('D',e_f),('E',prio_f),('G',stat_f),('H',c_f)])

make_sheet(wb, '\U0001F4E5 One-Time Tasks',
    ['Task Title *','Assignee Email(s)','Approver Email','Priority','Due Date','Client Name','Est. Hours','Description'],
    [28,32,32,10,14,20,10,28],
    [['Task - replace me',t1e,p1e,'high','2025-07-10',c1,'2',''],
     ['Task - replace me',t2e,'','medium','2025-07-15',c2,'1','']],
    [('B',e_f),('C',e_f),('D',prio_f),('F',c_f)])

make_sheet(wb, '\U0001F501 Recurring Tasks',
    ['Task Title *','Frequency *','Assignee Email(s)','Approver Email','Priority','Project Name','Start Date','Description'],
    [28,13,32,32,10,22,14,28],
    [['Weekly standup','weekly',t1e,'','medium','Project Alpha','2025-07-07',''],
     ['Monthly review','monthly',t2e,p1e,'high','Project Beta','2025-07-01','']],
    [('B',freq_f),('C',e_f),('D',e_f),('E',prio_f)])

if ca_mode:
    make_sheet(wb, '\U0001F3DB CA Compliance',
        ['Compliance Task Type *','Client Name','Assignee Email','Approver Email','Due Date','Priority','Frequency'],
        [30,22,32,32,14,10,14],
        [['GSTR 3B (Monthly)',c1,t1e,p1e,'2025-07-20','high','monthly'],
         ['TDS 26Q Return',c2,t2e,p1e,'2025-07-15','high','quarterly'],
         ['ITR Filing',c1,t1e,'','2025-07-31','urgent','']],
        [('B',c_f),('C',e_f),('D',e_f),('F',prio_f),('G',freq_f)])

with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
    tmp_path = tmp.name
wb.save(tmp_path)
with open(tmp_path, 'rb') as f:
    sys.stdout.buffer.write(base64.b64encode(f.read()))
os.unlink(tmp_path)
